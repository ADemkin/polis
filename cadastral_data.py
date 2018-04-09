from openpyxl import load_workbook
import json

LOCAL = 'cadastral_data.json'
DEPLOY = '/var/tmp/cadastral_data.json'

data_file = DEPLOY

def load_xlsx_data(xlsx_file):
    xls_data = load_workbook(xlsx_file)
    sheet_name = xls_data.sheetnames[0]
    sheet = xls_data[sheet_name]
    
    rows = sheet.rows
    all_data = []

    
    # todo: find coloumns by correct headers
    next(rows)  # skip headers
    for row in rows:
        cadastral_number = row[1].value  # Кадастровый номер
        # object = row[2].value # Объект. Находится только формула.
        project_name = row[3].value  # Проект
        project_glorax_competitor = row[4].value  # Конкурент проекта Glorax
        obj_type = row[5].value  # Тип
        if cadastral_number and obj_type:
            current = dict()
            current['cadastral_number'] = cadastral_number
            current['project_name'] = project_name
            current['project_glorax_competitor'] = project_glorax_competitor
            current['object_type'] = obj_type
            all_data.append(current)
    
    return json.dumps(all_data, indent=1, ensure_ascii=False)


def store_json_data(json_data, config_file=data_file):
    with open(config_file, 'wb') as file:
        file.write(json_data.encode())


def get_data_by_cadastral_number(cadastal_number, config_file=data_file):
    '''
    returns dict with all types for given cadastral number
    '''
    try:
        with open(config_file, 'rb') as file:
            json_data = json.load(file)
    except FileNotFoundError:
        return []
    data = [i for i in json_data if cadastal_number in i['cadastral_number']]
    if data:
        return data[0]
    else:
        return []


def test():
    file = "tasks/EGRN.xlsx"
    store_json_data(load_xlsx_data(file))
    # print(get_data_by_cadastral_number("78:32:0007503:31") == ['апарт, жилое'])
    # print(get_object_type_by_cadastral_number("78:14:0007519:4095") == ['жилое'])
    # print(get_object_type_by_cadastral_number("78:14:0007519:40195") == [])
    
    print(get_data_by_cadastral_number("78:32:0007503:31")['object_type'] == 'апарт, жилое')
    print(get_data_by_cadastral_number("78:14:0007519:4095")['object_type'] == 'жилое')
    
    print(get_data_by_cadastral_number("78:06:0002089:14")['project_glorax_competitor'] == 'Golden City')
    print(get_data_by_cadastral_number("78:06:0002089:1413") == [])
    
    print(get_data_by_cadastral_number("90:25:050801:263")['project_name'] == 'Gorizont Plaza')
    
    print(get_data_by_cadastral_number("90:25:010123:202"))



if __name__ == "__main__":
    test()
